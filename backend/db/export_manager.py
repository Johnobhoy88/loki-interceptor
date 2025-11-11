"""
Data export manager for CSV, JSON, and Excel exports
Enterprise-grade export capabilities with progress tracking
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from enum import Enum

from sqlalchemy.orm import Session

from backend.db.models import (
    Document, Validation, ValidationResult,
    Correction, AuditTrail, ExportLog
)
from backend.db.session import get_session


class ExportFormat(str, Enum):
    """Export format enumeration"""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    JSONL = "jsonl"  # JSON Lines format


class ExportManager:
    """
    Manages data exports in various formats
    Tracks export history and manages export files
    """

    def __init__(self, export_dir: str = 'data/exports'):
        """
        Initialize export manager

        Args:
            export_dir: Directory to store export files
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_documents(
        self,
        format: ExportFormat,
        client_id: str,
        requested_by: str,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        include_content: bool = True
    ) -> Dict[str, Any]:
        """
        Export documents to specified format

        Args:
            format: Export format
            client_id: Client identifier
            requested_by: User requesting export
            filters: Optional filters to apply
            fields: Optional list of fields to include
            include_content: Whether to include document content

        Returns:
            Export result information
        """
        with get_session() as session:
            # Create export log entry
            export_log = ExportLog(
                export_type=format.value,
                entity_type='document',
                filters=filters,
                fields=fields,
                requested_by=requested_by,
                client_id=client_id,
                status='pending'
            )
            session.add(export_log)
            session.flush()

            try:
                # Query documents
                query = session.query(Document).filter(
                    Document.client_id == client_id,
                    Document.is_deleted == False
                )

                # Apply filters
                if filters:
                    if 'document_type' in filters:
                        query = query.filter(Document.document_type == filters['document_type'])
                    if 'start_date' in filters:
                        query = query.filter(Document.created_at >= filters['start_date'])
                    if 'end_date' in filters:
                        query = query.filter(Document.created_at <= filters['end_date'])

                documents = query.all()

                # Convert to dictionaries
                data = [
                    doc.to_dict(include_content=include_content)
                    for doc in documents
                ]

                # Filter fields if specified
                if fields:
                    data = [
                        {k: v for k, v in row.items() if k in fields}
                        for row in data
                    ]

                # Export to file
                filename = f"documents_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
                file_path = self._export_data(filename, format, data)

                # Update export log
                file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                export_log.record_count = len(data)
                export_log.file_size_bytes = file_size
                export_log.file_path = str(file_path)
                export_log.status = 'completed'
                export_log.completed_at = datetime.utcnow()

                session.commit()

                return {
                    'export_id': export_log.id,
                    'file_path': str(file_path),
                    'record_count': len(data),
                    'file_size': file_size,
                    'format': format.value
                }

            except Exception as e:
                export_log.status = 'failed'
                export_log.error_message = str(e)
                session.commit()
                raise

    def export_validations(
        self,
        format: ExportFormat,
        client_id: str,
        requested_by: str,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        include_results: bool = False
    ) -> Dict[str, Any]:
        """
        Export validations to specified format

        Args:
            format: Export format
            client_id: Client identifier
            requested_by: User requesting export
            filters: Optional filters to apply
            fields: Optional list of fields to include
            include_results: Whether to include validation results

        Returns:
            Export result information
        """
        with get_session() as session:
            # Create export log entry
            export_log = ExportLog(
                export_type=format.value,
                entity_type='validation',
                filters=filters,
                fields=fields,
                requested_by=requested_by,
                client_id=client_id,
                status='pending'
            )
            session.add(export_log)
            session.flush()

            try:
                # Query validations
                query = session.query(Validation).filter(
                    Validation.client_id == client_id
                )

                # Apply filters
                if filters:
                    if 'status' in filters:
                        query = query.filter(Validation.status == filters['status'])
                    if 'risk_level' in filters:
                        query = query.filter(Validation.overall_risk == filters['risk_level'])
                    if 'start_date' in filters:
                        query = query.filter(Validation.created_at >= filters['start_date'])
                    if 'end_date' in filters:
                        query = query.filter(Validation.created_at <= filters['end_date'])

                validations = query.all()

                # Convert to dictionaries
                data = []
                for validation in validations:
                    row = validation.to_dict()

                    # Include results if requested
                    if include_results:
                        results = session.query(ValidationResult).filter(
                            ValidationResult.validation_id == validation.id
                        ).all()
                        row['results'] = [r.to_dict() for r in results]

                    data.append(row)

                # Filter fields if specified
                if fields and not include_results:
                    data = [
                        {k: v for k, v in row.items() if k in fields}
                        for row in data
                    ]

                # Export to file
                filename = f"validations_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
                file_path = self._export_data(filename, format, data)

                # Update export log
                file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                export_log.record_count = len(data)
                export_log.file_size_bytes = file_size
                export_log.file_path = str(file_path)
                export_log.status = 'completed'
                export_log.completed_at = datetime.utcnow()

                session.commit()

                return {
                    'export_id': export_log.id,
                    'file_path': str(file_path),
                    'record_count': len(data),
                    'file_size': file_size,
                    'format': format.value
                }

            except Exception as e:
                export_log.status = 'failed'
                export_log.error_message = str(e)
                session.commit()
                raise

    def export_audit_trail(
        self,
        format: ExportFormat,
        client_id: str,
        requested_by: str,
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Export audit trail to specified format

        Args:
            format: Export format
            client_id: Client identifier
            requested_by: User requesting export
            filters: Optional filters to apply
            fields: Optional list of fields to include

        Returns:
            Export result information
        """
        with get_session() as session:
            # Create export log entry
            export_log = ExportLog(
                export_type=format.value,
                entity_type='audit',
                filters=filters,
                fields=fields,
                requested_by=requested_by,
                client_id=client_id,
                status='pending'
            )
            session.add(export_log)
            session.flush()

            try:
                # Query audit entries
                query = session.query(AuditTrail).filter(
                    AuditTrail.client_id == client_id
                )

                # Apply filters
                if filters:
                    if 'action' in filters:
                        query = query.filter(AuditTrail.action == filters['action'])
                    if 'entity_type' in filters:
                        query = query.filter(AuditTrail.entity_type == filters['entity_type'])
                    if 'actor_id' in filters:
                        query = query.filter(AuditTrail.actor_id == filters['actor_id'])
                    if 'start_date' in filters:
                        query = query.filter(AuditTrail.timestamp >= filters['start_date'])
                    if 'end_date' in filters:
                        query = query.filter(AuditTrail.timestamp <= filters['end_date'])

                audits = query.all()

                # Convert to dictionaries
                data = [audit.to_dict() for audit in audits]

                # Filter fields if specified
                if fields:
                    data = [
                        {k: v for k, v in row.items() if k in fields}
                        for row in data
                    ]

                # Export to file
                filename = f"audit_trail_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
                file_path = self._export_data(filename, format, data)

                # Update export log
                file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                export_log.record_count = len(data)
                export_log.file_size_bytes = file_size
                export_log.file_path = str(file_path)
                export_log.status = 'completed'
                export_log.completed_at = datetime.utcnow()

                session.commit()

                return {
                    'export_id': export_log.id,
                    'file_path': str(file_path),
                    'record_count': len(data),
                    'file_size': file_size,
                    'format': format.value
                }

            except Exception as e:
                export_log.status = 'failed'
                export_log.error_message = str(e)
                session.commit()
                raise

    def _export_data(
        self,
        filename: str,
        format: ExportFormat,
        data: List[Dict[str, Any]]
    ) -> Path:
        """
        Export data to file in specified format

        Args:
            filename: Base filename (without extension)
            format: Export format
            data: Data to export

        Returns:
            Path to exported file
        """
        if format == ExportFormat.CSV:
            return self._export_csv(filename, data)
        elif format == ExportFormat.JSON:
            return self._export_json(filename, data)
        elif format == ExportFormat.JSONL:
            return self._export_jsonl(filename, data)
        elif format == ExportFormat.EXCEL:
            return self._export_excel(filename, data)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    def _export_csv(self, filename: str, data: List[Dict[str, Any]]) -> Path:
        """
        Export data to CSV

        Args:
            filename: Base filename
            data: Data to export

        Returns:
            Path to CSV file
        """
        file_path = self.export_dir / f"{filename}.csv"

        if not data:
            # Create empty file
            file_path.touch()
            return file_path

        # Get all unique keys from all records
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(fieldnames)

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for row in data:
                # Convert complex types to strings
                clean_row = {}
                for key, value in row.items():
                    if isinstance(value, (dict, list)):
                        clean_row[key] = json.dumps(value)
                    elif isinstance(value, datetime):
                        clean_row[key] = value.isoformat()
                    else:
                        clean_row[key] = value
                writer.writerow(clean_row)

        return file_path

    def _export_json(self, filename: str, data: List[Dict[str, Any]]) -> Path:
        """
        Export data to JSON

        Args:
            filename: Base filename
            data: Data to export

        Returns:
            Path to JSON file
        """
        file_path = self.export_dir / f"{filename}.json"

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        return file_path

    def _export_jsonl(self, filename: str, data: List[Dict[str, Any]]) -> Path:
        """
        Export data to JSON Lines format

        Args:
            filename: Base filename
            data: Data to export

        Returns:
            Path to JSONL file
        """
        file_path = self.export_dir / f"{filename}.jsonl"

        with open(file_path, 'w', encoding='utf-8') as f:
            for row in data:
                f.write(json.dumps(row, default=str) + '\n')

        return file_path

    def _export_excel(self, filename: str, data: List[Dict[str, Any]]) -> Path:
        """
        Export data to Excel

        Args:
            filename: Base filename
            data: Data to export

        Returns:
            Path to Excel file
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        file_path = self.export_dir / f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        if not data:
            wb.save(file_path)
            return file_path

        # Get all unique keys
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(fieldnames)

        # Write header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = row_data.get(field)

                # Convert complex types
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                elif isinstance(value, datetime):
                    value = value.isoformat()

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        return file_path

    def get_export_history(
        self,
        client_id: str,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """
        Get export history for a client

        Args:
            client_id: Client identifier
            limit: Maximum results

        Returns:
            List of export log entries
        """
        with get_session() as session:
            exports = session.query(ExportLog).filter(
                ExportLog.client_id == client_id
            ).order_by(ExportLog.requested_at.desc()).limit(limit).all()

            return [export.to_dict() for export in exports]

    def delete_export(self, export_id: int) -> bool:
        """
        Delete an export file and log entry

        Args:
            export_id: Export log ID

        Returns:
            True if deleted, False otherwise
        """
        with get_session() as session:
            export_log = session.query(ExportLog).filter(
                ExportLog.id == export_id
            ).first()

            if not export_log:
                return False

            # Delete file if exists
            if export_log.file_path and os.path.exists(export_log.file_path):
                os.remove(export_log.file_path)

            # Delete log entry
            session.delete(export_log)
            session.commit()

            return True

    def cleanup_expired_exports(self, retention_days: int = 30) -> int:
        """
        Clean up expired export files

        Args:
            retention_days: Days to retain exports

        Returns:
            Number of deleted exports
        """
        from datetime import timedelta

        cutoff = datetime.utcnow() - timedelta(days=retention_days)
        deleted_count = 0

        with get_session() as session:
            old_exports = session.query(ExportLog).filter(
                ExportLog.requested_at < cutoff
            ).all()

            for export_log in old_exports:
                # Delete file if exists
                if export_log.file_path and os.path.exists(export_log.file_path):
                    os.remove(export_log.file_path)

                session.delete(export_log)
                deleted_count += 1

            session.commit()

        return deleted_count
